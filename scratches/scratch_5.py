# 03/01/2019

from openpyxl import load_workbook #Load the original file
from openpyxl import Workbook  #Create and Save excel file

class File_Transfer():
    def __init__(self,file_path):
        self.file_path = 'c:/CSV/CIQ-DSR 7.0.1-2-22-2019.xlsx'

    #To read the input from original file and create workbook,Creating a LIST with each tabs.
    def read_CIQ(self,file_path):
        wb = load_workbook(self.file_path)
        self.ws_list = wb.sheetnames
        for name in ws_list:
            print(f"here is the name of each list : {name}!")
        return

    #To create the target CSV file as target_file (with only 1 sheet)
    def create_csv(self,target_file_path,target_file_name,target_sheet_name):
        self.target_file_path = target_file_path
        self.target_file_name = target_file_name
        self.target_sheet_name = target_sheet_name

        target_wb = Workbook()
        #changing sheet name
        target_ws = target_wb.active
        target_ws.title = f"{target_sheet_name}"

        target_wb.save(target_file_name)



    #This is to get the first Row from each tab(sheet) of input file and make it as first Colum in target file
    def take_column_make_row(self,sheet,column_list):












yue = File_Transfer('c:/CSV/CIQ-DSR 7.0.1-2-22-2019.xlsx')

print (yue.read_csv('c:/CSV/CIQ-DSR 7.0.1-2-22-2019.xlsx'))




